import openpyxl

try:
    wb = openpyxl.load_workbook("BUGS IN IQAF HTML.xlsx")
    sheet = wb.active
    print("Sheet Title:", sheet.title)
    print("Max Row:", sheet.max_row)
    print("Max Col:", sheet.max_column)
    
    for r in range(1, sheet.max_row + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
        if any(row_vals):
            print(f"Row {r}: {row_vals}")
except Exception as e:
    print("Error reading excel file:", e)

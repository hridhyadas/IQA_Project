import openpyxl

wb = openpyxl.load_workbook("BUGS IN IQAF HTML.xlsx")
sheet = wb.active

with open("scratch/bugs_list.txt", "w", encoding="utf-8") as f:
    f.write(f"Sheet Title: {sheet.title}\n")
    f.write(f"Max Row: {sheet.max_row}\n")
    f.write(f"Max Col: {sheet.max_column}\n\n")
    
    for r in range(1, sheet.max_row + 1):
        row_vals = [sheet.cell(row=r, column=c).value for c in range(1, sheet.max_column + 1)]
        if any(row_vals):
            # Format row nicely
            vals_str = ", ".join([str(v) for v in row_vals if v is not None])
            f.write(f"Row {r}: {vals_str}\n")
print("Done writing bugs_list.txt")

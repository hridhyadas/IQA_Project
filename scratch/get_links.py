import openpyxl

wb = openpyxl.load_workbook("BUGS IN IQAF HTML.xlsx")
sheet = wb.active

for r in [23, 24, 25, 26, 27, 28]:
    c1 = sheet.cell(row=r, column=1)
    c2 = sheet.cell(row=r, column=2)
    c3 = sheet.cell(row=r, column=3)
    
    url1 = c1.hyperlink.target if c1.hyperlink else None
    url2 = c2.hyperlink.target if c2.hyperlink else None
    
    print(f"Row {r}:")
    print(f"  Desc: {c3.value}")
    print(f"  Figma Link: {url1}")
    print(f"  Design Link: {url2}")
